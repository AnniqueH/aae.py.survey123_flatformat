# -------------------------------------------------------------------------------
# Name:        process_survey123_datafile
# Purpose:
#
# Author:      Adrian Kitchingman
#
# Created:     13/05/2022
# Copyright:   (c) ak34 2022
# Licence:     <your licence>

# Last Update: 08/05/2023

# ~~~~ TO DO ~~~~~~~~~~~~
#   Add timestamp
#   Add name of input file
#   Error Check output file not open
#   Navigate to input file
#       https://stackoverflow.com/questions/9319317/quick-and-easy-file-dialog-in-python
# -------------------------------------------------------------------------------

# import copy
# import csv
import os
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
import process_survey123_field_data_classes as cls
import process_survey123_field_data_functions as func
import local_vars as localvars
from tkinter import *
import tkinter.messagebox
from tkinter.filedialog import askopenfilename #as fd

root = Tk()
root.withdraw()
root.update()

io_path = localvars.io_path
filename = askopenfilename(initialdir=io_path, title="Open Survey123 XLSX File")
in_xlfile = os.path.basename(filename)
io_path = filename.replace(in_xlfile, '')

while True:   # repeat until the try statement succeeds
    try:
##        myfile = open(filename, "r+") # or "a+", whatever you need
        workbook = load_workbook(filename)
        break                             # exit the loop
    except IOError:
##        input("Could not open file! Please close Excel. Press Enter to retry.")
        answer = tkinter.messagebox.askokcancel("Open File Error", "Could not open file! Please close Excel. Press OK to retry.")
        # restart the loop


##root.mainloop()
root.destroy()

print(workbook.sheetnames)

# #['Fish_Survey_XVEFMAPX_0', 'site_location_repeat_1', 'shot_repeat_2', 'observed_fish_repeat_3', 'fish_sample_repeat_4']

# # ## READ SITE SURVEY DATA =======================================================================================

survey_list = func.read_in_excel_tab(workbook.worksheets[0])

# # ## READ location DATA =======================================================================================

loc_list = func.read_in_excel_tab(workbook['site_location_repeat_1'])

# # ## READ SHOT DATA =======================================================================================

shot_list = func.read_in_excel_tab(workbook['shot_repeat_2'])

# # ## READ observed DATA =======================================================================================

obs_list = func.read_in_excel_tab(workbook['observed_fish_repeat_3'])

# # ## READ sampled DATA =======================================================================================

sample_list = func.read_in_excel_tab(workbook['fish_sample_repeat_4'])


# #### -- ### Sort the samples so any defined shots are at the top.
# # -- ## t[16] + " " + t[2][::-1] + " " + t[4] + " " + t[5]
sample_list.sort(key=lambda x: 0 if x[2] is None else int(x[2]), reverse=True)
sample_list.sort(key=lambda x: x[16])

print('Building data arrays...')

# #loop through site survey sheet
# #    get site survey info
for svy in survey_list:
    site_id = svy[1]
    site_code = svy[2]
    survey_date = svy[3]
    # survey_date = "{0}".format(svy[3])
    # survey_date = datetime.strptime(survey_date, '%Y-%m-%d %H:%M:%S')
    # survey_date = survey_date.strftime('%d/%m/%Y')
    gear_type = svy[4]
    personnel1 = svy[5]
    personnel2 = svy[6]
    depth_secchi = svy[7]
    depth_max = svy[8]
    depth_avg = svy[9]
    section_condition = svy[10]
    time_start = svy[11]
    time_end = svy[12]
    project_name = svy[13]
    survey_notes = svy[14]
    site_survey_ts = svy[15]
    water_qual_depth = svy[16]
    ec_25c = svy[17]
    water_temp = svy[18]
    do_mgl = svy[19]
    do_perc = svy[20]
    ph = svy[21]
    turbidity_ntu = svy[22]
    chlorophyll = svy[23]
    creationdate = svy[25]
    creator = svy[26]
    editdate = svy[27]
    editor = svy[28]
    data_x = svy[29]
    data_y = svy[30]
    x_start = 0.0
    y_start = 0.0
    x_finish = 0.0
    y_finish = 0.0

    # filter locations and loop through to populate start finish coordinates
    site_locations = list(filter(lambda x: x[3] == site_id, loc_list))
    if site_locations is not None:
        for coord in site_locations:
            if coord[2] == 'site_start' and x_start == 0.0:
                x_start = coord[8]
                y_start = coord[9]
            elif coord[2] == 'site_finish' and x_finish == 0.0:
                x_finish = coord[8]
                y_finish = coord[9]

# #    loop through shot sheet
# #        get shot info into array
    site_shots = list(filter(lambda x: x[11] == site_id, shot_list))
    if len(site_shots) > 0:

        for sht in site_shots:

            section_number = '1' if sht[2] is None and len(site_shots) == 1 else sht[2]
            # section_number = sht[2]
            shot_id = sht[1]
            electro_seconds = sht[3]
            soak_minutes_per_unit = sht[4]
            section_time_start = sht[5]
            section_time_end = sht[6]
            volts = sht[7]
            amps = sht[8]
            pulses_per_second = sht[9]
            percent_duty_cycle = sht[10]

            func.site_survey_info.append(cls.SiteSurvey(site_id, site_code, survey_date, gear_type, personnel1, personnel2, depth_secchi, depth_max, depth_avg, section_condition, time_start, time_end, project_name, survey_notes, water_qual_depth, ec_25c, water_temp, do_mgl, do_perc, ph, turbidity_ntu, chlorophyll, creationdate, creator, editdate, editor, data_x, data_y, x_start, y_start, x_finish, y_finish, shot_id, section_number, electro_seconds, soak_minutes_per_unit, section_time_start, section_time_end, volts, amps, pulses_per_second, percent_duty_cycle))

            # obs_fish = 0
            shot_obs = list(filter(lambda x: x[8] == shot_id and x[7] is not None, obs_list))
            # print('site: {0} shot: {1} shot_obs count: {2}'.format(site_id, shot_id, len(shot_obs)))

            if len(shot_obs) > 0:
                for ob in shot_obs:
                    obs_id = ob[1]

                    if ob[2] is not None or ob[3] is not None or ob[4] is not None:
                        if ob[2] is None:
                            species = ob[4]
                        else:
                            if ob[2].find('Non Listed Species') >= 0:
                                species = ob[3]
                            else:
                                species = ob[2]

                        collected = ob[5]
                        observed = ob[6]
                        if collected is None:
                            collected = 0
                        if observed is None:
                            observed = 0
                        if collected != 0 or observed != 0:
                            func.sssoc_info.append(cls.SiteObs(site_id, section_number, species, collected, observed, collected, shot_id, obs_id))
                            # obs_fish += 1
                    # else:
                    #     if len(shot_obs) == 1:
                    #         # # If no obs hit but shot exists then include shot
                    #         site_samples = list(filter(lambda x: x[16] == site_id, sample_list))

            else:
                # # If no obs hit but shot exists then include shot
                site_samples = list(filter(lambda x: x[16] == site_id, sample_list))
                # if there are samples then add empty shot
                # print('site: {0} samples: {1}'.format(site_id, len(site_samples)))

                if section_condition == 'yes' and len(site_samples) > 0:
                    # print('got here site: {0}, shot: {1}, obs_count: {2}, samp count: {3}'.format(site_code, section_number,
                    #                                                                      len(shot_obs),
                    #                                                                      len(site_samples)))
                    func.sssoc_info.append(cls.SiteObs(site_id, section_number, '', 0, 0, 0, shot_id, ''))
                    # if site_id == '2c3c5e53-ff04-4fa1-9637-e67331d77b50': print('got here')
                # no samples then add 'No Fish'
                elif section_condition == 'yes' and len(site_samples) == 0:
                    func.sssoc_info.append(cls.SiteObs(site_id, section_number, 'No Fish', 0, 0, 0, shot_id, ''))
                # #                print('add shot')
    else:
        # # if no shots exist then put in 1 shot if fishable or samples are present and add site info
        site_samples = list(filter(lambda x: x[16] == site_id, sample_list))
        if section_condition == 'yes' and len(site_samples) > 0:
            # print('samples present {0} - {1}'.format(site_id, len(site_samples)))
            func.site_survey_info.append(cls.SiteSurvey(site_id, site_code, survey_date, gear_type, personnel1, personnel2, depth_secchi, depth_max, depth_avg, section_condition, time_start, time_end, project_name, survey_notes, water_qual_depth, ec_25c, water_temp, do_mgl, do_perc, ph, turbidity_ntu, chlorophyll, creationdate, creator, editdate, editor, x_start, y_start, x_finish, y_finish, '', '1', electro_seconds, soak_minutes_per_unit, section_time_start, section_time_end, volts, amps, pulses_per_second, percent_duty_cycle))
            func.sssoc_info.append(cls.SiteObs(site_id, '1', '', 0, 0, 0, '', ''))

        elif section_condition == 'yes' and len(site_samples) == 0:
            # print('no samples present {0}'.format(site_id))
            func.site_survey_info.append(cls.SiteSurvey(site_id, site_code, survey_date, gear_type, personnel1, personnel2, depth_secchi, depth_max, depth_avg, section_condition, time_start, time_end, project_name, survey_notes, water_qual_depth, ec_25c, water_temp, do_mgl, do_perc, ph, turbidity_ntu, chlorophyll, creationdate, creator, editdate, editor, x_start, y_start, x_finish, y_finish, '', '1', electro_seconds, soak_minutes_per_unit, section_time_start, section_time_end, volts, amps, pulses_per_second, percent_duty_cycle))
            func.sssoc_info.append(cls.SiteObs(site_id, '1', 'No Fish', 0, 0, 0, '', ''))


# #for i in sssoc_info:
# #    print('site: {0}: {1}'.format(i.site_id, i.species))

print('Processing data into new format...')

test = 0

if test == 1:
    # #    for i in sssoc_info:
    # #        print('{0},{1},{2},{3},{4},{5},{6},{7}'.format(i[0], i[1],i[2],i[3],i[4],i[5],i[6],i[7]))

    # #        for s in sample_list:
    # #            print('{0},{2},{3},{4},{5},{6},{7},{8},{9},{10},{11},{12},{13},{14},{15},{16},'.format(s[0],s[1],s[2],s[3],s[4],s[5],s[6],s[7],s[8],s[9],s[10],s[11],s[12],s[13],s[14],s[15],s[16]))

    for i in func.site_survey_info:
        print('{0}, {1}, {2}, {3}, {4}, {5}, {6}, {7}, {8}, {9}, {10}, {11}, {12}, {13}, {14}, {15}, {16}, {17}, {18}, {19}, {20}, {21}, {22}, {23}, {24}, {25}, {26}, {27}, {28}, {29}, {30}, {31}, {32}, {33}, {34}'.format(i['k_site_id'], i['k_site_code'], i['k_survey_date'], i['k_gear_type'], i['k_personnel1'], i['k_personnel2'], i['k_depth_secchi'], i['k_depth_max'], i['k_depth_avg'], i['k_section_condition'], i['k_survey_notes'], i['k_water_qual_depth'], i['k_ec_25c'], i['k_water_temp'], i['k_do_mgl'], i['k_do_perc'], i['k_ph'], i['k_turbidity_ntu'], i['k_chlorophyll'], i['k_creationdate'], i['k_creator'], i['k_editdate'], i['k_editor'], i['k_x'], i['k_y'], i['k_shot_id'], i['k_section_number'], i['k_electro_seconds'], i['k_soak_minutes_per_unit'], i['k_time_start'], i['k_time_end'], i['k_volts'], i['k_amps'], i['k_pulses_per_second'], i['k_percent_duty_cycle']))

else:
    # #loop through sampled
    # #    fw = open('O:/DATA/DOCS/PROJECTS/VEFMAP/FIELD_DATA/test_out.csv', 'w')
    # #    fw.write('site_code,sample_site_id,shot,species,coll,fl,tl,w,over\n')

    xl_header = list(("project",
                        "site_code",
                        "x_coordinate",
                        "y_coordinate",
                        "finish_x_coordinate",
                        "finish_y_coordinate",
                        "survey_date",
                        "gear_type",
                        "personnel1",
                        "personnel2",
                        "depth_secchi",
                        "depth_max",
                        "depth_avg",
                        "section_condition",
                        "time_start",
                        "time_end",
                        "survey_notes",
                        "section_number",
                        "electro_seconds",
                        "soak_minutes_per_unit",
                        "section_time_start",
                        "section_time_end",
                        "volts",
                        "amps",
                        "pulses_per_second",
                        "percent_duty_cycle",
                        "species",
                        "fork_length",
                        "total_length",
                        "weight",
                        "collected",
                        "observed",
                        "recapture",
                        "pit",
                        "external_tag_no",
                        "genetics_label",
                        "otoliths_label",
                        "fauna_notes",
                        "water_qual_depth",
                        "ec_25c",
                        "water_temp",
                        "do_mgl",
                        "do_perc",
                        "ph",
                        "turbidity_ntu",
                        "chlorophyll",
                        "Site_GlobalID",
                        "Shot_GlobalID",
                        "Obs_GlobalID",
                        "Sample_GlobalID",
                        "data_recording_x",
                        "data_recording_y"))

    wb = openpyxl.Workbook()
    ws_write = wb.active
    ws_write.title = "Raw Data"
    ws2_write = wb.create_sheet("Tally Results", 1)

    row_count = 1
    func.write_row(ws_write, row_count, 1, xl_header)

    row_count += 1

    prev_sample_site_id = ''

    shots_used = []

    samples_present = False
    for smp in sample_list:

        # #        objectid=smp[0]
        # #        globalid=smp[1]
        # #        section_number_samp=smp[2]
        # #        species_samp_custom=smp[3]
        # #        species_samp=smp[4]
        # #        fork_length=smp[5]
        # #        total_length=smp[6]
        # #        weight=smp[7]
        # #        coll=smp[8]

        # print(smp[0])

        if smp[3] is not None or smp[4] is not None:
            samples_present = True
            sample_id = smp[1]
            section_number = 0 if smp[2] is None else smp[2]
            # #        if smp[2] is not None:
            # #            print(smp[2])
            species = smp[4] if smp[3] is None else smp[3]
            fl = '' if smp[5] is None else smp[5]
            tl = '' if smp[6] is None else smp[6]
            w = '' if smp[7] is None else smp[7]

            skip_samp = 0
            if (smp[8] is None or smp[8] == 0) and fl == '' and tl == '' and w == '':
                skip_samp = 1

            if skip_samp == 0:

                coll = 1 if smp[8] is None or smp[8] == 0 else smp[8]
                recapture = smp[9]
                pit = "{0}".format(smp[11]) if smp[11] is not None else ''
                external_tag_no = smp[10]
                genetics_label = smp[13]
                otoliths_label = smp[14]
                fauna_notes = smp[15]
                sample_site_id = smp[16]

                if sample_site_id != prev_sample_site_id:
                    # #            print(sample_site_id ,prev_sample_site_id)
                    if prev_sample_site_id != '':
                        # # OUTPUT EXTRA DATA
                        row_count = func.extra_record_output(ws_write, prev_sample_site_id, row_count)
                    prev_sample_site_id = sample_site_id
                    # #-----get site and shot info
                    # #-------get species
                    # #
                    # #- -----------GET RANDOM SHOT
                    # #        print(sample_site_id, species, sample_id, section_number)


                shot_i = func.get_random_shot(sample_site_id, species) if section_number == 0 else str(section_number)

                #get site/survey for the shot selected
                sub_site_survey_info = list(filter(lambda x: x['k_site_id'] == sample_site_id and x['k_section_number'] == shot_i, func.site_survey_info))

                extra_shot = 0
                if len(sub_site_survey_info) == 0:
                    extra_shot = 1
                    shot_i = str(section_number)
                    sub_site_survey_info = list(filter(lambda x: x['k_site_id'] == sample_site_id and x['k_section_number'] == '1', func.site_survey_info))
                    print('*** NO SHOT ERROR for\nsite: {0}\nsample:{1}\nshot: {2}\n*** -----------'.format(sample_site_id, sample_id, shot_i))

                if len(sub_site_survey_info) > 1:
                    print('*** MULTIPLE SURVEY ERROR for\nsite: {0}\nshot: {1}\n*** --------------'.format(sample_site_id, shot_i))

                # #        shot_i = s[1]
                if isinstance(shot_i, str):
                    shot_i = int(shot_i)

                shots_used.append(shot_i)

                for ss_row in sub_site_survey_info:
                    func.adjust_species_shot(sample_site_id, species, str(shot_i), coll)
                    gear_type = ss_row.gear_type
                    section_time_start = ss_row.section_time_start
                    section_time_end = ss_row.section_time_end
                    electro_seconds = ss_row.electro_seconds
                    volts = ss_row.volts
                    amps = ss_row.amps
                    pulses_per_second = ss_row.pulses_per_second
                    percent_duty_cycle = ss_row.percent_duty_cycle

                    if extra_shot == 1:
                        ss_row.gear_type = 'EXTRA_SHOT_IN_SAMPLES'
                        ss_row.section_time_start = ''
                        ss_row.section_time_end = ''
                        ss_row.electro_seconds = 0
                        ss_row.volts = ''
                        ss_row.amps = ''
                        ss_row.pulses_per_second = ''
                        ss_row.percent_duty_cycle = ''

                    func.write_excel_row(ws_write, row_count, ss_row, shot_i, species, fl, tl, w, coll, 0, recapture, pit, external_tag_no, genetics_label, otoliths_label, fauna_notes, '', sample_id)

                    ss_row.gear_type = gear_type
                    ss_row.section_time_start = section_time_start
                    ss_row.section_time_end = section_time_end
                    ss_row.electro_seconds = electro_seconds
                    ss_row.volts = volts
                    ss_row.amps = amps
                    ss_row.pulses_per_second = pulses_per_second
                    ss_row.percent_duty_cycle = percent_duty_cycle

                    row_count += 1


    # if samples_present == False:

# #    fw.close()

# #####################################################################################################################
# #####################################################################################################################
# #####################################################################################################################
# #####################################################################################################################
# #####################################################################################################################
# #####################################################################################################################

# ##### FINISH UP ########################################
    # # OUTPUT EXTRA DATA FOR LAST SITE
    row_count = func.extra_record_output(ws_write, prev_sample_site_id, row_count)

    # ADD any no samples shots (but other shots in site had fish)
    for sobs in func.sssoc_info:
        if sobs[1] not in shots_used and sobs[2] == '' and sobs[7] == '':
            row_count = func.extra_record_output_no_fish_shot(ws_write, sobs[0], sobs[1], row_count)

    # ADD any no samples fish sites
    prev_site_id = ''
    for sobs in func.sssoc_info:
        # print(sobs[0])
        if sobs[0] != prev_site_id:
            site_obs = list(filter(lambda x: x[0] == sobs[0], func.sssoc_info))
            site_samples = list(filter(lambda x: x[16] == sobs[0], sample_list))
            sample_flag = False
            if len(site_samples) == 0:
                # print('{0},{1},{2},{3},{4},{5},{6},{7}'.format(sobs[0],sobs[1],sobs[2],sobs[3],sobs[4],sobs[5],sobs[6],sobs[7]))
                row_count = func.extra_record_output(ws_write, sobs[0], row_count)
                sample_flag = True

        prev_site_id = sobs[0]

    # ###----- ORDER BY site, date, shot, species, observed, collected
    # ###----- sheet_sort_rows(ws, row_start, row_end=0, cols=None, sorter=None, reverse=False)
    func.sheet_sort_rows(ws_write, 2, 0, [2, 7, 18, 27, 32, 31])
    func.set_col_date_style(ws_write, (7-1))

# #    sub_sssoc_info = list(filter(lambda x: x[5] != 0, sssoc_info))
# #    if len(sub_sssoc_info) > 0:
    row_count = 1
    xl_header = list(("site_id", "section_number", "species", "collected", "observed", "collected_tally", "shot_id", "obs_id"))
    func.write_row(ws2_write, row_count, 1, xl_header)

    for i in func.sssoc_info:
        xl_row = list((i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]))
        row_count += 1
        func.write_row(ws2_write, row_count, 1, xl_row)


    dnow = datetime.now()
    fdt = dnow.strftime("%y") + dnow.strftime("%m") + dnow.strftime("%d")


    out_xlfile = in_xlfile.replace('(', '').replace(')', '')
    out_xlfile = out_xlfile.replace(".xlsx", "_FlatFormat_" + fdt + ".xlsx")

    wb.save(io_path + out_xlfile)

    print('\nFormatting complete. New Excel file is at:\n{0}\n\n'.format(io_path + out_xlfile))

# #    print('{0},{1},{2},{3},{4},{5},{6},{7}'.format(i[0], i[1],i[2],i[3],i[4],i[5],i[6],i[7]))

# #####################################################################################################################
# #####################################################################################################################
# #####################################################################################################################
# #####################################################################################################################
# #####################################################################################################################
# #####################################################################################################################
# #####################################################################################################################
# #####################################################################################################################
# #####################################################################################################################
# #####################################################################################################################
# #####################################################################################################################




